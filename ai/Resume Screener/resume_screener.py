import os
import re
import json
import time
import logging
import traceback
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple, Union
import concurrent.futures
import pandas as pd
import base64
from io import BytesIO

# PDF extraction libraries
import PyPDF2
import pdfplumber
import pdf2image

# DOCX extraction
import docx2txt

# LangChain imports
from langchain_core.prompts import ChatPromptTemplate
from langchain_google_genai import ChatGoogleGenerativeAI
from langchain.output_parsers import PydanticOutputParser
from langchain_core.pydantic_v1 import BaseModel, Field
from langchain_core.messages import HumanMessage

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[
        logging.StreamHandler(),  # Log to console
        logging.FileHandler("resume_screener.log")  # Also log to file
    ]
)
logger = logging.getLogger("resume_screener")

# Models for structured output
class SkillMatch(BaseModel):
    """Representation of a skill match between resume and job description"""
    skill_name: str = Field(description="Name of the skill")
    present_in_resume: bool = Field(description="Whether the skill is present in the resume")
    importance_level: int = Field(description="Importance level from 1-5, with 5 being critical")
    proficiency_level: Optional[int] = Field(None, description="Estimated proficiency level 1-5 based on context")
    context: Optional[str] = Field(None, description="Context from resume showing skill usage")

class ExperienceMatch(BaseModel):
    """Representation of experience match between resume and job description"""
    area: str = Field(description="Area of experience (e.g., 'Python development', 'Project management')")
    years_required: float = Field(description="Years of experience required in job description")
    years_present: float = Field(description="Years of experience present in resume")
    relevance_score: int = Field(description="How relevant the experience is to the job, 1-5 scale")
    context: Optional[str] = Field(None, description="Context from resume showing the experience")

class EducationMatch(BaseModel):
    """Representation of education match between resume and job description"""
    requirement: str = Field(description="Education requirement from job description")
    has_match: bool = Field(description="Whether candidate meets education requirement")
    details: str = Field(description="Details about candidate's education")

class ResumeAnalysis(BaseModel):
    """Complete analysis of a resume compared to job description"""
    candidate_name: str = Field(description="Name of the candidate")
    overall_fit_score: int = Field(description="Overall fit score from 0-100")
    skill_matches: List[SkillMatch] = Field(description="Analysis of skill matches")
    experience_matches: List[ExperienceMatch] = Field(description="Analysis of experience matches")
    education_match: EducationMatch = Field(description="Analysis of education match")
    strengths: List[str] = Field(description="Key strengths of the candidate relative to the position")
    gaps: List[str] = Field(description="Key gaps or areas for improvement")
    summary: str = Field(description="Brief summary explaining the fit score")

class ResumeExtractor:
    """Extract text content from resume files of various formats"""
    
    def __init__(self, model=None):
        """Initialize the extractor
        
        Args:
            model: The Gemini model instance to use for image-based extraction
        """
        self.model = model
    
    def extract_text(self, file_path: str) -> Dict[str, Any]:
        """Extract text content from a resume file
        
        Args:
            file_path: Path to the resume file
            
        Returns:
            Dictionary containing the extracted text and metadata
        """
        file_path = Path(file_path)
        if not file_path.exists():
            logger.error(f"File not found: {file_path}")
            raise FileNotFoundError(f"File not found: {file_path}")
        
        file_extension = file_path.suffix.lower()
        
        try:
            logger.info(f"Extracting text from {file_path} (format: {file_extension})")
            if file_extension == '.pdf':
                return self._extract_from_pdf(file_path)
            elif file_extension in ['.docx', '.doc']:
                return self._extract_from_docx(file_path)
            elif file_extension in ['.txt', '.text']:
                return self._extract_from_txt(file_path)
            else:
                logger.error(f"Unsupported file format: {file_extension}")
                raise ValueError(f"Unsupported file format: {file_extension}")
        except Exception as e:
            logger.error(f"Error extracting text from {file_path}: {str(e)}")
            logger.error(traceback.format_exc())
            raise
    
    def _extract_from_pdf(self, file_path: Path) -> Dict[str, Any]:
        """Extract text from PDF file
        
        Args:
            file_path: Path to the PDF file
            
        Returns:
            Dictionary with extracted text and metadata
        """
        # Try PyPDF2 first (faster but less accurate for complex PDFs)
        try:
            with open(file_path, 'rb') as f:
                pdf_reader = PyPDF2.PdfReader(f)
                text = ""
                for page_num in range(len(pdf_reader.pages)):
                    text += pdf_reader.pages[page_num].extract_text() + "\n"
                
                # If we got substantial text, return it
                if len(text.strip()) > 100:
                    logger.info(f"Successfully extracted text using PyPDF2: {len(text)} chars")
                    return {
                        "text": text,
                        "metadata": {
                            "source": file_path.name,
                            "pages": len(pdf_reader.pages),
                            "extractor": "PyPDF2"
                        }
                    }
        except Exception as e:
            logger.warning(f"PyPDF2 extraction failed for {file_path}: {str(e)}")
        
        # If PyPDF2 failed or returned little text, try pdfplumber
        try:
            with pdfplumber.open(file_path) as pdf:
                text = ""
                for page in pdf.pages:
                    text += page.extract_text() + "\n"
                
                # If we got substantial text, return it
                if len(text.strip()) > 100:
                    logger.info(f"Successfully extracted text using pdfplumber: {len(text)} chars")
                    return {
                        "text": text,
                        "metadata": {
                            "source": file_path.name,
                            "pages": len(pdf.pages),
                            "extractor": "pdfplumber"
                        }
                    }
        except Exception as e:
            logger.warning(f"pdfplumber extraction failed for {file_path}: {str(e)}")
        
        # If both failed or return little text and Gemini model is available, use it
        if self.model:
            try:
                logger.info(f"Using Gemini model for image-based extraction of {file_path}")
                
                # Convert PDF pages to images
                try:
                    images = pdf2image.convert_from_path(file_path)
                    logger.info(f"Successfully converted PDF to {len(images)} images")
                except Exception as e:
                    logger.error(f"Failed to convert PDF to images: {str(e)}")
                    logger.error(traceback.format_exc())
                    raise
                
                # Extract text from each page using Gemini
                text = ""
                for i, image in enumerate(images):
                    logger.info(f"Processing page {i+1}/{len(images)} with Gemini")
                    # Convert image to base64
                    buffered = BytesIO()
                    image.save(buffered, format="JPEG")
                    img_str = base64.b64encode(buffered.getvalue()).decode()
                    
                    # Create message with image
                    message = [
                        {
                            "type": "text",
                            "text": "Extract all text from this image of a resume page. Include ALL text visible in the image, presented as plain text."
                        },
                        {
                            "type": "image_url",
                            "image_url": {
                                "url": f"data:image/jpeg;base64,{img_str}"
                            }
                        }
                    ]
                    
                    # Get response from Gemini
                    try:
                        human_message = HumanMessage(content=message)
                        response = self.model.invoke([human_message])
                        text += response.content + "\n\n"
                        logger.info(f"Successfully extracted text from page {i+1} using Gemini")
                    except Exception as e:
                        logger.error(f"Error with Gemini for page {i+1}: {str(e)}")
                        logger.error(traceback.format_exc())
                        raise
                
                logger.info(f"Successfully extracted text using Gemini: {len(text)} chars")
                return {
                    "text": text,
                    "metadata": {
                        "source": file_path.name,
                        "pages": len(images),
                        "extractor": "gemini-vision"
                    }
                }
            except Exception as e:
                logger.error(f"Gemini extraction failed for {file_path}: {str(e)}")
                logger.error(traceback.format_exc())
                raise
        
        # If we get here, all extraction methods failed
        logger.error(f"All extraction methods failed for PDF: {file_path}")
        raise ValueError(f"Could not extract text from PDF: {file_path}")
    
    def _extract_from_docx(self, file_path: Path) -> Dict[str, Any]:
        """Extract text from DOCX file
        
        Args:
            file_path: Path to the DOCX file
            
        Returns:
            Dictionary with extracted text and metadata
        """
        try:
            text = docx2txt.process(file_path)
            logger.info(f"Successfully extracted text from DOCX: {len(text)} chars")
            return {
                "text": text,
                "metadata": {
                    "source": file_path.name,
                    "extractor": "docx2txt"
                }
            }
        except Exception as e:
            logger.error(f"Error extracting text from DOCX {file_path}: {str(e)}")
            logger.error(traceback.format_exc())
            raise
    
    def _extract_from_txt(self, file_path: Path) -> Dict[str, Any]:
        """Extract text from text file
        
        Args:
            file_path: Path to the text file
            
        Returns:
            Dictionary with extracted text and metadata
        """
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                text = f.read()
            
            logger.info(f"Successfully extracted text from TXT: {len(text)} chars")
            return {
                "text": text,
                "metadata": {
                    "source": file_path.name,
                    "extractor": "plain-text"
                }
            }
        except Exception as e:
            logger.error(f"Error extracting text from TXT {file_path}: {str(e)}")
            logger.error(traceback.format_exc())
            raise
    
    def extract_candidate_name(self, text: str) -> str:
        """Attempt to extract candidate name from resume text
        
        Args:
            text: Full text of the resume
            
        Returns:
            Extracted name or "Unknown Candidate"
        """
        try:
            # Use the first few lines which typically contain the name
            first_lines = text.split('\n')[:10]
            first_text = ' '.join(first_lines)
            
            # Simple heuristic: Look for a name pattern (First Last) near the top
            name_pattern = re.compile(r'\b([A-Z][a-z]+)\s+([A-Z][a-z]+)\b')
            match = name_pattern.search(first_text)
            if match:
                name = match.group(0)
                logger.info(f"Extracted candidate name: {name}")
                return name
            
            # If no clear name pattern, use a more generic approach
            words = first_text.split()
            for i in range(len(words) - 1):
                if (words[i][0].isupper() and words[i].isalpha() and 
                    words[i+1][0].isupper() and words[i+1].isalpha()):
                    name = f"{words[i]} {words[i+1]}"
                    logger.info(f"Extracted candidate name (fallback): {name}")
                    return name
            
            logger.warning("Could not extract candidate name")
            return "Unknown Candidate"
        except Exception as e:
            logger.error(f"Error extracting candidate name: {str(e)}")
            return "Unknown Candidate"

class ResumeAnalyzer:
    """Analyze resumes against job descriptions using LLM"""
    
    def __init__(self, model_name: str = "gemini-2.5-flash-lite", api_key: Optional[str] = None, temperature: float = 0.2):
        """Initialize the analyzer
        
        Args:
            model_name: Name of the Gemini model to use
            api_key: Google API key (or reads from env var GOOGLE_API_KEY)
            temperature: Temperature setting for the model
        """
        logger.info(f"Initializing ResumeAnalyzer with model: {model_name}")
        self.model_name = model_name
        
        if api_key:
            os.environ["GOOGLE_API_KEY"] = api_key
            logger.info("Using provided API key")
        
        self.api_key = os.environ.get("GOOGLE_API_KEY")
        if not self.api_key:
            logger.warning("GOOGLE_API_KEY not found. Analyzer will not function correctly.")
        else:
            logger.info(f"Using API key (first 4 chars): {self.api_key[:4]}***")
        
        try:
            # Initialize the model
            self.llm = ChatGoogleGenerativeAI(
                model=model_name,
                temperature=temperature,
                google_api_key=self.api_key
            )
            logger.info("Successfully initialized LLM")
            
            # Test the model with a simple query to verify API key works
            test_response = self.llm.invoke("Hello, please respond with 'API working'")
            logger.info(f"Test response: {test_response.content[:20]}...")
            
            # Pass the model to the extractor for image-based extraction
            self.extractor = ResumeExtractor(model=self.llm)
            self._setup_llm_chain()
        except Exception as e:
            logger.error(f"Error initializing LLM: {str(e)}")
            logger.error(traceback.format_exc())
            raise
    
    def _setup_llm_chain(self):
        """Set up the LLM chain for resume analysis"""
        try:
            # Create output parser
            self.output_parser = PydanticOutputParser(pydantic_object=ResumeAnalysis)
            
            # Create prompt template
            self.prompt = ChatPromptTemplate.from_template("""
                You are an expert resume screener tasked with evaluating candidate resumes against a specific job description.
                Analyze the following resume against the job description to determine how well the candidate fits the role.
                
                JOB DESCRIPTION:
                {job_description}
                
                RESUME:
                {resume_text}
                
                Provide a detailed, objective analysis including:
                1. The candidate's name
                2. An overall fit score from 0-100
                3. Matching of skills required vs. present in resume
                4. Matching of experience required vs. present in resume
                5. Matching of education requirements
                6. Key strengths relative to the position
                7. Key gaps or areas for improvement
                8. A brief summary explaining the fit score
                
                Use a data-driven approach and avoid bias in your evaluation.
                Format your response according to the following schema:
                
                {format_instructions}
                
                Focus on relevance, depth, and specificity of experience rather than just keyword matching.
                """
            )
            
            # Create chain
            self.chain = self.prompt | self.llm | self.output_parser
            logger.info("Successfully set up LLM chain")
        except Exception as e:
            logger.error(f"Error setting up LLM chain: {str(e)}")
            logger.error(traceback.format_exc())
            raise
    
    def analyze_resume(self, resume_path: str, job_description: str) -> ResumeAnalysis:
        """Analyze a single resume against a job description
        
        Args:
            resume_path: Path to the resume file
            job_description: Text of the job description
            
        Returns:
            Structured analysis of the resume
        """
        logger.info(f"Analyzing resume: {resume_path}")
        
        # Extract text from resume
        try:
            extracted = self.extractor.extract_text(resume_path)
            resume_text = extracted["text"]
            logger.info(f"Successfully extracted resume text ({len(resume_text)} chars)")
            
            # Extract candidate name
            candidate_name = self.extractor.extract_candidate_name(resume_text)
        except Exception as e:
            logger.error(f"Error extracting text from {resume_path}: {str(e)}")
            logger.error(traceback.format_exc())
            raise
        
        # Prepare inputs
        try:
            inputs = {
                "job_description": job_description,
                "resume_text": resume_text,
                "format_instructions": self.output_parser.get_format_instructions()
            }
            
            # Log sizes for debugging
            logger.info(f"Job description size: {len(job_description)} chars")
            logger.info(f"Resume text size: {len(resume_text)} chars")
            logger.info(f"Format instructions size: {len(inputs['format_instructions'])} chars")
            
            # Analyze with LLM
            logger.info("Starting LLM analysis...")
            analysis = self.chain.invoke(inputs)
            
            # If name wasn't extracted well, use our heuristic
            if analysis.candidate_name == "Unknown" or not analysis.candidate_name:
                logger.info(f"Replacing unknown name with extracted name: {candidate_name}")
                analysis.candidate_name = candidate_name
            
            logger.info(f"Analysis complete. Score: {analysis.overall_fit_score}")
            return analysis
        except Exception as e:
            logger.error(f"Error analyzing resume with LLM: {str(e)}")
            logger.error(traceback.format_exc())
            raise
    
    def analyze_resume_batch(self, resume_paths: List[str], job_description: str, 
                            max_workers: int = 4) -> List[ResumeAnalysis]:
        """Analyze multiple resumes against a job description
        
        Args:
            resume_paths: List of paths to resume files
            job_description: Text of the job description
            max_workers: Maximum number of parallel workers
            
        Returns:
            List of structured analyses for each resume
        """
        logger.info(f"Batch analyzing {len(resume_paths)} resumes")
        
        results = []
        
        # Process in parallel
        with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
            future_to_resume = {
                executor.submit(self.analyze_resume, resume_path, job_description): resume_path
                for resume_path in resume_paths
            }
            
            for future in concurrent.futures.as_completed(future_to_resume):
                resume_path = future_to_resume[future]
                try:
                    analysis = future.result()
                    results.append(analysis)
                    logger.info(f"Completed analysis for {resume_path}: Score {analysis.overall_fit_score}")
                except Exception as e:
                    logger.error(f"Analysis failed for {resume_path}: {str(e)}")
                    logger.error(traceback.format_exc())
        
        # Sort by fit score descending
        results.sort(key=lambda x: x.overall_fit_score, reverse=True)
        return results
    
    def export_results_to_csv(self, analyses: List[ResumeAnalysis], output_path: str) -> None:
        """Export analysis results to CSV file
        
        Args:
            analyses: List of resume analyses
            output_path: Path to save the CSV file
        """
        data = []
        for analysis in analyses:
            # Basic info
            row = {
                "Candidate Name": analysis.candidate_name,
                "Overall Fit Score": analysis.overall_fit_score,
                "Summary": analysis.summary
            }
            
            # Skills
            skills_present = [skill.skill_name for skill in analysis.skill_matches if skill.present_in_resume]
            skills_missing = [skill.skill_name for skill in analysis.skill_matches if not skill.present_in_resume]
            
            row["Skills Present"] = ", ".join(skills_present)
            row["Skills Missing"] = ", ".join(skills_missing)
            
            # Experience
            exp_summary = []
            for exp in analysis.experience_matches:
                if exp.years_present >= exp.years_required:
                    status = "Sufficient"
                else:
                    status = f"Insufficient ({exp.years_present}/{exp.years_required} yrs)"
                exp_summary.append(f"{exp.area}: {status}")
            
            row["Experience Summary"] = "; ".join(exp_summary)
            
            # Education
            row["Education Match"] = "Yes" if analysis.education_match.has_match else "No"
            row["Education Details"] = analysis.education_match.details
            
            # Strengths and gaps
            row["Strengths"] = "; ".join(analysis.strengths)
            row["Gaps"] = "; ".join(analysis.gaps)
            
            data.append(row)
        
        # Create DataFrame and export
        df = pd.DataFrame(data)
        df.to_csv(output_path, index=False)
        logger.info(f"Results exported to {output_path}")
        
        return df
    
    def export_results_to_excel(self, analyses: List[ResumeAnalysis], output_path: str) -> None:
        """Export analysis results to Excel file with formatting
        
        Args:
            analyses: List of resume analyses
            output_path: Path to save the Excel file
        """
        # Create DataFrame
        data = []
        for analysis in analyses:
            # Basic info
            row = {
                "Candidate Name": analysis.candidate_name,
                "Overall Fit Score": analysis.overall_fit_score,
                "Summary": analysis.summary
            }
            
            # Skills
            skills_present = [skill.skill_name for skill in analysis.skill_matches if skill.present_in_resume]
            skills_missing = [skill.skill_name for skill in analysis.skill_matches if not skill.present_in_resume]
            
            row["Skills Present"] = ", ".join(skills_present)
            row["Skills Missing"] = ", ".join(skills_missing)
            
            # Experience
            exp_summary = []
            for exp in analysis.experience_matches:
                if exp.years_present >= exp.years_required:
                    status = "Sufficient"
                else:
                    status = f"Insufficient ({exp.years_present}/{exp.years_required} yrs)"
                exp_summary.append(f"{exp.area}: {status}")
            
            row["Experience Summary"] = "; ".join(exp_summary)
            
            # Education
            row["Education Match"] = "Yes" if analysis.education_match.has_match else "No"
            row["Education Details"] = analysis.education_match.details
            
            # Strengths and gaps
            row["Strengths"] = "; ".join(analysis.strengths)
            row["Gaps"] = "; ".join(analysis.gaps)
            
            data.append(row)
        
        # Create DataFrame
        df = pd.DataFrame(data)
        
        # Export to Excel with formatting
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Resume Analysis"
            
            # Write header
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == 1:  # Header row
                        cell.fill = header_fill
                        cell.font = header_font
            
            # Auto-adjust column width
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = max(max_length + 2, 10)  # Add padding and set minimum width
                ws.column_dimensions[column_letter].width = min(adjusted_width, 50)  # Cap width
            
            # Conditional formatting for scores
            for row_idx, row in enumerate(df.iterrows(), 2):  # Start from row 2 (after header)
                score_cell = ws.cell(row=row_idx, column=3)  # Assuming score is in column 3
                
                # Convert score to integer for comparison (fix for the bug)
                try:
                    score = int(score_cell.value)
                    
                    if score >= 80:
                        score_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif score >= 60:
                        score_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    else:
                        score_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                except (ValueError, TypeError):
                    # If score can't be converted to int, skip coloring
                    logger.warning(f"Could not convert score '{score_cell.value}' to integer for Excel formatting")
                    pass
            
            # Save workbook
            wb.save(output_path)
            logger.info(f"Results exported to Excel file: {output_path}")
        except ImportError:
            # If openpyxl is not installed, fall back to basic export
            logger.warning("openpyxl not installed, exporting without formatting")
            df.to_excel(output_path, index=False)
            logger.info(f"Results exported to {output_path} (without formatting)")
        
        return df

class ResumeScreener:
    """Main class for resume screening functionality"""
    
    def __init__(self, model_name: str = "gemini-2.5-flash-lite", api_key: Optional[str] = None,
                max_workers: int = 4, temperature: float = 0.2):
        """Initialize the resume screener
        
        Args:
            model_name: Name of the Gemini model to use
            api_key: Google API key (or reads from env var GOOGLE_API_KEY)
            max_workers: Maximum number of parallel workers for batch processing
            temperature: Temperature setting for the model
        """
        logger.info(f"Initializing ResumeScreener with model: {model_name}")
        try:
            self.analyzer = ResumeAnalyzer(model_name=model_name, api_key=api_key, temperature=temperature)
            self.max_workers = max_workers
            logger.info("ResumeScreener initialized successfully")
        except Exception as e:
            logger.error(f"Failed to initialize ResumeScreener: {str(e)}")
            logger.error(traceback.format_exc())
            raise
    
    def screen_single_resume(self, resume_path: str, job_description: str) -> Dict[str, Any]:
        """Screen a single resume against a job description
        
        Args:
            resume_path: Path to the resume file
            job_description: Text of the job description
            
        Returns:
            Dictionary with analysis results
        """
        logger.info(f"Screening single resume: {resume_path}")
        try:
            analysis = self.analyzer.analyze_resume(resume_path, job_description)
            
            # Convert to dictionary for easier serialization
            result = analysis.dict()
            logger.info(f"Screening complete. Score: {result['overall_fit_score']}")
            
            return result
        except Exception as e:
            logger.error(f"Error screening resume {resume_path}: {str(e)}")
            logger.error(traceback.format_exc())
            raise
    
    def screen_multiple_resumes(self, resume_dir_or_paths: Union[str, List[str]], 
                               job_description: str, output_format: str = 'excel',
                               output_path: Optional[str] = None) -> Dict[str, Any]:
        """Screen multiple resumes against a job description
        
        Args:
            resume_dir_or_paths: Directory containing resumes or list of file paths
            job_description: Text of the job description
            output_format: Format to save results ('csv', 'excel', or 'json')
            output_path: Path to save results (optional)
            
        Returns:
            Dictionary with summary results and path to output file
        """
        # Get list of resume paths
        try:
            if isinstance(resume_dir_or_paths, str) and os.path.isdir(resume_dir_or_paths):
                # It's a directory, get all resume files
                resume_dir = resume_dir_or_paths
                resume_paths = []
                for file in os.listdir(resume_dir):
                    ext = os.path.splitext(file)[1].lower()
                    if ext in ['.pdf', '.docx', '.doc', '.txt']:
                        resume_paths.append(os.path.join(resume_dir, file))
                logger.info(f"Found {len(resume_paths)} resume files in directory: {resume_dir}")
            elif isinstance(resume_dir_or_paths, list):
                # It's a list of paths
                resume_paths = resume_dir_or_paths
                logger.info(f"Processing {len(resume_paths)} resume files from list")
            else:
                error_msg = "resume_dir_or_paths must be a directory or a list of file paths"
                logger.error(error_msg)
                raise ValueError(error_msg)
            
            if not resume_paths:
                error_msg = "No resume files found"
                logger.error(error_msg)
                raise ValueError(error_msg)
        except Exception as e:
            logger.error(f"Error finding resume files: {str(e)}")
            logger.error(traceback.format_exc())
            raise
        
        # Analyze resumes
        try:
            logger.info(f"Starting batch analysis of {len(resume_paths)} resumes")
            start_time = time.time()
            analyses = self.analyzer.analyze_resume_batch(
                resume_paths=resume_paths,
                job_description=job_description,
                max_workers=self.max_workers
            )
            processing_time = time.time() - start_time
            logger.info(f"Batch analysis completed in {processing_time:.2f} seconds")
        except Exception as e:
            logger.error(f"Error in batch analysis: {str(e)}")
            logger.error(traceback.format_exc())
            raise
        
        # Generate default output path if not provided
        if not output_path:
            timestamp = time.strftime("%Y%m%d_%H%M%S")
            if output_format == 'csv':
                output_path = f"resume_analysis_{timestamp}.csv"
            elif output_format == 'excel':
                output_path = f"resume_analysis_{timestamp}.xlsx"
            else:  # json
                output_path = f"resume_analysis_{timestamp}.json"
            logger.info(f"Using default output path: {output_path}")
        
        # Export results
        try:
            logger.info(f"Exporting results in {output_format} format to {output_path}")
            if output_format == 'csv':
                self.analyzer.export_results_to_csv(analyses, output_path)
            elif output_format == 'excel':
                self.analyzer.export_results_to_excel(analyses, output_path)
            else:  # json
                with open(output_path, 'w') as f:
                    json.dump([analysis.dict() for analysis in analyses], f, indent=2)
            logger.info(f"Results exported to {output_path}")
        except Exception as e:
            logger.error(f"Error exporting results: {str(e)}")
            logger.error(traceback.format_exc())
            raise
        
        # Prepare summary
        summary = {
            "total_resumes": len(resume_paths),
            "successfully_analyzed": len(analyses),
            "processing_time_seconds": processing_time,
            "average_score": sum(a.overall_fit_score for a in analyses) / len(analyses) if analyses else 0,
            "top_candidate": analyses[0].candidate_name if analyses else None,
            "top_score": analyses[0].overall_fit_score if analyses else None,
            "output_file": output_path
        }
        
        logger.info(f"Batch processing summary: {summary}")
        return summary
    
    def explain_analysis(self, analysis: Union[ResumeAnalysis, Dict[str, Any]]) -> str:
        """Generate a human-readable explanation of the analysis results
        
        Args:
            analysis: ResumeAnalysis object or dictionary
            
        Returns:
            Formatted string explaining the analysis
        """
        # Convert dict to ResumeAnalysis if needed
        if isinstance(analysis, dict):
            try:
                analysis = ResumeAnalysis(**analysis)
            except Exception as e:
                logger.error(f"Error converting dict to ResumeAnalysis: {str(e)}")
                logger.error(traceback.format_exc())
                raise
        
        explanation = f"""
RESUME ANALYSIS FOR: {analysis.candidate_name}

OVERALL FIT SCORE: {analysis.overall_fit_score}/100

SUMMARY:
{analysis.summary}

SKILLS ASSESSMENT:
"""
        
        # Add skills
        for skill in analysis.skill_matches:
            status = "✓" if skill.present_in_resume else "✗"
            importance = "★" * skill.importance_level
            explanation += f"  {status} {skill.skill_name} {importance}\n"
            if skill.present_in_resume and skill.context:
                explanation += f"     Context: {skill.context}\n"
        
        # Add experience
        explanation += "\nEXPERIENCE ASSESSMENT:\n"
        for exp in analysis.experience_matches:
            if exp.years_present >= exp.years_required:
                status = "✓"
            else:
                status = "✗"
            explanation += f"  {status} {exp.area}: {exp.years_present} yrs vs. {exp.years_required} yrs required\n"
            if exp.context:
                explanation += f"     Context: {exp.context}\n"
        
        # Add education
        explanation += "\nEDUCATION ASSESSMENT:\n"
        status = "✓" if analysis.education_match.has_match else "✗"
        explanation += f"  {status} {analysis.education_match.requirement}\n"
        explanation += f"     Details: {analysis.education_match.details}\n"
        
        # Add strengths and gaps
        explanation += "\nKEY STRENGTHS:\n"
        for strength in analysis.strengths:
            explanation += f"  • {strength}\n"
        
        explanation += "\nAREAS FOR IMPROVEMENT:\n"
        for gap in analysis.gaps:
            explanation += f"  • {gap}\n"
        
        return explanation


# Example usage
if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description="Smart Resume Screener")
    parser.add_argument("--resume", help="Path to resume file or directory containing resumes")
    parser.add_argument("--job-description", help="Path to job description file")
    parser.add_argument("--output", help="Path to output file")
    parser.add_argument("--format", choices=["csv", "excel", "json"], default="excel", 
                       help="Output format (default: excel)")
    parser.add_argument("--model", default="gemini-2.5-flash", 
                       help="Name of the Gemini model to use")
    parser.add_argument("--api-key", help="Google API key")
    parser.add_argument("--max-workers", type=int, default=4, 
                       help="Maximum number of parallel workers for batch processing")
    
    args = parser.parse_args()
    
    # Check for required arguments
    if not args.resume:
        parser.error("--resume is required")
    if not args.job_description:
        parser.error("--job-description is required")
    
    # Read job description
    with open(args.job_description, 'r', encoding='utf-8') as f:
        job_description = f.read()
    
    # Initialize screener
    screener = ResumeScreener(
        model_name=args.model,
        api_key=args.api_key,
        max_workers=args.max_workers
    )
    
    # Process resumes
    if os.path.isdir(args.resume):
        # Batch processing
        summary = screener.screen_multiple_resumes(
            resume_dir_or_paths=args.resume,
            job_description=job_description,
            output_format=args.format,
            output_path=args.output
        )
        
        print("\nBatch Processing Summary:")
        print(f"Total Resumes: {summary['total_resumes']}")
        print(f"Successfully Analyzed: {summary['successfully_analyzed']}")
        print(f"Processing Time: {summary['processing_time_seconds']:.2f} seconds")
        print(f"Average Score: {summary['average_score']:.2f}")
        print(f"Top Candidate: {summary['top_candidate']} (Score: {summary['top_score']})")
        print(f"Results saved to: {summary['output_file']}")
    else:
        # Single resume
        analysis = screener.screen_single_resume(args.resume, job_description)
        
        # Print explanation
        explanation = screener.explain_analysis(analysis)
        print(explanation)
        
        # Save results if output path provided
        if args.output:
            if args.format == 'json':
                with open(args.output, 'w') as f:
                    json.dump(analysis, f, indent=2)
            else:
                # Convert to DataFrame
                df = pd.DataFrame([{
                    "Candidate Name": analysis["candidate_name"],
                    "Overall Fit Score": analysis["overall_fit_score"],
                    "Summary": analysis["summary"],
                    "Skills Present": ", ".join(s["skill_name"] for s in analysis["skill_matches"] 
                                              if s["present_in_resume"]),
                    "Skills Missing": ", ".join(s["skill_name"] for s in analysis["skill_matches"] 
                                              if not s["present_in_resume"]),
                    "Experience Summary": "; ".join(f"{e['area']}: {e['years_present']}/{e['years_required']} yrs" 
                                                  for e in analysis["experience_matches"]),
                    "Education Match": "Yes" if analysis["education_match"]["has_match"] else "No",
                    "Education Details": analysis["education_match"]["details"],
                    "Strengths": "; ".join(analysis["strengths"]),
                    "Gaps": "; ".join(analysis["gaps"])
                }])
                
                if args.format == 'csv':
                    df.to_csv(args.output, index=False)
                else:  # excel
                    df.to_excel(args.output, index=False)
            
            print(f"\nResults saved to: {args.output}")
